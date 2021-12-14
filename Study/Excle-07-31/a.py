from openpyxl import load_workbook
import PySimpleGUI as sg


def load_table(choose):
    wb = load_workbook(choose)
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append(row)
    for x in range(1, len(rows)):
        data_text = []
        # 订单编号
        order = str(rows[x][1].value)
        # 电话1
        phone = str(rows[x][7].value)
        # 邮政编码
        postal_code = str(rows[x][9].value)
        # 所属城市
        city =  str(rows[x][10].value)
        # 邮寄地址1(完整导出)
        mail_address_first = str(rows[x][11].value)
        # 邮寄地址2
        mail_address_second = str(rows[x][12].value)
        data_text.append(order)
        data_text.append(phone)
        data_text.append(postal_code)
        data_text.append(city)
        data_text.append(mail_address_first)
        data_text.append(mail_address_second)
        yield data_text

def dispose_list(list_data):
    for index, value in enumerate(list_data):
        # print(index, value)
        for number in range(0, len(list_data)):
            if index == 5 and number == 0:
                continue
            elif index == 5 and number == 1:
                continue
            elif index == 5 and number == 2:
                continue
            else:
                if number != index:
                    if value in list_data[number]:
                        return list_data[0]

# 选取文件
def pySimpleGUI_choose_file():
    file_address = sg.popup_get_file('请选择你要读取的表格文件：')
    return file_address

# choose = r'E:\插件\cod核单\20核单.xlsx'
results = load_table(pySimpleGUI_choose_file())
with open('重复.txt', 'w', encoding='utf-8') as f:
    for result in results:
        error_order = dispose_list(result)
        try:
            f.write(error_order + '\n')
        except TypeError:
            continue
