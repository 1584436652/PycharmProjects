'''
@Time    : 2021/10/13 14:08
@Author  : LKT
@FileName: whatshop.py
@Software: PyCharm
 
'''
import PySimpleGUI as sg
import time

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook


class WhatShopPhone(object):

    def __init__(self):
        options = webdriver.ChromeOptions()
        options.add_argument('disable-infobars')
        options.add_experimental_option("detach", True)
        self.driver = webdriver.Chrome("D:\Download\chromedriver_win32\chromedriver.exe", options=options)
        self.driver.maximize_window()
        self.wait = WebDriverWait(self.driver, 3)
        self.driver.get('https://web.whatsapp.com/')
        self.phone_url = 'https://api.whatsapp.com/send?phone={}'

    @property
    def pySimpleGUI_choose_file(self):
        file_address = sg.popup_get_file('请选择你要读取的表格文件：')
        return file_address

    def read_phone(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        rows = []
        for row in ws.iter_rows():
            rows.append(row)
        for x in range(1, len(rows)):
            data_text = []
            phone = str(rows[x][0].value)
            message = str(rows[x][1].value)
            nation = str(rows[x][2].value)
            # print(phone, message)
            data_text.append(phone)
            data_text.append(message)
            data_text.append(nation)
            yield data_text

    def verbal_trick(self, file_name):
        with open(f'{file_name}.txt', 'r', encoding='utf-8') as fp:
            return fp.read()

    def running(self, phone, message):
        try:
            self.driver.get(self.phone_url.format(phone))
            time.sleep(1.5)
            # 继续对话
            self.wait.until(
                EC.element_to_be_clickable((By.ID, "action-button"))).click()
            time.sleep(2)
            # 使用网页版
            self.wait.until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="fallback_block"]//a[@class="_36or"]'))).click()
            time.sleep(1)
            try:
                time.sleep(2)
                # 信息输入框
                self.wait.until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//div[@id="main"]//div[@class="_13NKt copyable-text selectable-text"]'
                         ))).send_keys(message)
                time.sleep(1)
            except TimeoutException:
                print("Phone number shared via url is invalid.")
                return f'电话不存在 {phone}'
            else:
                # 发送信息
                self.wait.until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="main"]//span[@data-testid="send"]'))).click()
                time.sleep(2)
        except Exception:
            print('Network problem or error')
            return f'网络问题或执行错误 {phone}'

    def run(self):
        results = self.read_phone(self.pySimpleGUI_choose_file)
        with open('error_phone.txt', 'w', encoding='utf-8') as fp_phone:
            for result in results:
                verbal_str = self.verbal_trick(result[2])
                str_txt = f'Order number:{result[1]} {verbal_str}'
                fp_phone.write(self.running(result[0], str_txt) + '\n')

shop = WhatShopPhone()
shop.run()

