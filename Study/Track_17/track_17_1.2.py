'''
@Time    : 2021/10/12 12:48
@Author  : LKT
@FileName: track_17_1.1.py
@Software: PyCharm
 
'''
import requests
import json
import math
import time
import random

from openpyxl import load_workbook
from openpyxl import Workbook
from retry import retry

"""
17track查单
"""

class Track_17(object):

    def __init__(self, cookie=None):

        self.headers = {
            "user-agent":"",
            'Referer': 'https://t.17track.net/zh-cn',
            'Cookie': cookie
        }
        self.proxies = {
            # "https": "https://47.108.154.194:8080"
        }
        self.url = 'https://t.17track.net/restapi/track'

    headers_list = [
        "Mozilla/5.0 (Macintosh; U; Intel Mac OS X 10_6_8; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
        "Mozilla/5.0 (Windows; U; Windows NT 6.1; en-us) AppleWebKit/534.50 (KHTML, like Gecko) Version/5.1 Safari/534.50",
        "Mozilla/5.0 (Windows NT 10.0; WOW64; rv:38.0) Gecko/20100101 Firefox/38.0",
        "Mozilla/5.0 (Windows NT 10.0; WOW64; Trident/7.0; .NET4.0C; .NET4.0E; .NET CLR 2.0.50727; .NET CLR 3.0.30729; .NET CLR 3.5.30729; InfoPath.3; rv:11.0) like Gecko",
        "Mozilla/5.0 (compatible; MSIE 9.0; Windows NT 6.1; Trident/5.0)",
        "Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 6.0; Trident/4.0)",
        "Mozilla/4.0 (compatible; MSIE 7.0; Windows NT 6.0)",
        "Mozilla/4.0 (compatible; MSIE 6.0; Windows NT 5.1)",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.6; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
        "Mozilla/5.0 (Windows NT 6.1; rv:2.0.1) Gecko/20100101 Firefox/4.0.1",
        "Opera/9.80 (Macintosh; Intel Mac OS X 10.6.8; U; en) Presto/2.8.131 Version/11.11",
        "Opera/9.80 (Windows NT 6.1; U; en) Presto/2.8.131 Version/11.11",
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_7_0) AppleWebKit/535.11 (KHTML, like Gecko) Chrome/17.0.963.56 Safari/535.11",
        "Mozilla/6.0 (iPhone; CPU iPhone OS 8_0 like Mac OS X) AppleWebKit/536.26 (KHTML, like Gecko) Version/8.0 Mobile/10A5376e Safari/8536.25"
    ]

    @property
    def my_headers(self):
        user_agent = random.choice(self.headers_list)
        self.headers["user-agent"] = user_agent
        return self.headers

    # 读取跟踪号表格
    def read_trcak(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        rows = []
        tracking_number_list = []
        for row in ws.iter_rows():
            rows.append(row)
        for x in range(1, len(rows)):
            # 物流跟踪号
            tracking_number = str(rows[x][0].value)
            # message = str(rows[x][1].value)
            # nation = str(rows[x][2].value)
            tracking_number_list.append(tracking_number)
        return tracking_number_list

    # 构造data
    def trcak_data(self, track_list: list, *args):
        messages= []
        for order in track_list[args[0]:args[1]]:
            result = {"num":order,"fc":0,"sc":0}
            messages.append(result)
        data = {
            "data": messages,
            "guid":"",
            "timeZoneOffset":-480
        }
        return data

    @retry(delay=200)
    def make_response(self, url, headers, data):
        print(f'正在请求--{url}')
        # print(headers)
        response = requests.post(url=url, headers=headers, data=json.dumps(data), proxies=self.proxies)
        html_data = response.json()
        print(html_data)
        if html_data['dat']:
            return html_data
        print("您的查询过于频繁，请稍后查询，等待重新尝试中(200s)...")
        raise Exception

    # 检查数据是否正确, ["track"]["e"]是否有数据，为None，需要重新请求
    def check_json_data(self, track_none):
        orders = track_none["dat"]
        try:
            for order in orders:
                if order["track"]["e"]:
                    continue
            return track_none
        except Exception:
                return None

    # 解析json获取物流信息
    def track_parse(self, track_json):
        orders = track_json["dat"]
        for order in orders:
            # 跟踪号
            track_numbers = order["no"]
            # 物流状态
            delivery_status = str(order["track"]["e"])
            # print(track_numbers, delivery_status)
            # 物流详情
            delivery_details = order["track"]["z1"]
            track_append = []
            for delivery_detail in delivery_details:
                # 到达站点时间
                delivery_date = delivery_detail["a"]
                # 到达站点地址
                delivery_address = delivery_detail["z"]
                # print(delivery_date, delivery_address)
                track_append.append({delivery_date: delivery_address})
                # 返回跟踪后全部物流详情
            yield {
                track_numbers: [delivery_status, track_append]
            }

    # 修改解析后的物流状态
    def modify_status(self, messages: dict):
        status = {
            "0": "查询不到",
            "10": "运输途中",
            "35": "投递失败",
            "30": "到达待取",
            "40": "成功签收",
            "50": "可能异常",
            "60": "运输过久",
        }
        for message_key, massage_value in messages.items():
            # 把对应的物流状态改为中文
            messages[message_key][0] = status[messages[message_key][0]]
        # print(messages)
        return messages

    # 读取异常对应中文
    def abnormal_detail(self, filename):
        with open(f'{filename}.json', 'r', encoding='utf-8') as f:
            abnormal_json = f.read()
            abnormal_list = json.loads(abnormal_json)
            # print(abnormal_list)
            return abnormal_list

    # 用abnormal_detail相对应的可能异常改为中文
    def dispose(self, lists: list, dicts: dict):
        # 遍历传进的列表字典，匹配相对应的字段，保存到end_save_to
        end_save_to = []
        for i, j in dicts.items():
            end_save_to.append(i)
            end_save_to.append(j[0])
        for item in lists:
            for abnormal_key, abnormal_values in item.items():
                for address in dicts.values():
                    for titles in address[1]:
                        for date, title in titles.items():
                            if address[0] == "可能异常" or address[0] == "投递失败":
                                # 判断对应异常是否在地址详情里，查询到就改为中文
                                if abnormal_key in title:
                                   alter =  titles[date] = abnormal_values
                                   end_save_to.append(alter)
        # print(end_save_to)
        return end_save_to

    # 保存为excel, 传进[[]]
    def save_track_status(self, track_save: list):
        file_location = f'物流{self.date_name}.xlsx'
        wb = Workbook()
        ws = wb.active
        print(f'总条数：{len(track_save)}')
        ws.append(['track_number', '签收状态', '物流详情'])
        for j in range(2, len(track_save)+2):
            # 判断单个列表的长度，长度大于2有物流详情
            if len(track_save[j-2]) > 2:
                ws[f'A{j}'] = track_save[j-2][0]
                ws[f'B{j}'] = track_save[j-2][1]
                ws[f'C{j}'] = ';'.join(track_save[j-2][:1:-1])
            else:
                ws[f'A{j}'] = track_save[j-2][0]
                ws[f'B{j}'] = track_save[j-2][1]
        # ws.append(track_save)
        wb.save(file_location)
        print(f'文件名为:{file_location}')

    # 当前时间
    @property
    def date_name(self):
        return time.strftime("%Y%m%d%H%M%S", time.localtime(int(time.time())))

    # 跟踪号条数/40向上取整
    def ceil_int(self, totals):
        return math.ceil(totals / 40)

    def run(self):
        save_data = []
        track_list = track.read_trcak('track.xlsx')
        totals = self.ceil_int(len(track_list))
        abnormal = self.abnormal_detail("异常列表")
        # 一次只能查找40条
        start_number = 0
        end_number = 40
        while totals >= 1:
            data = self.trcak_data(track_list, start_number, end_number)
            track_json = self.make_response(self.url, self.my_headers,  data)
            check = self.check_json_data(track_json)
            if check != None:
                print(f"{start_number}-{end_number}物流详情获取成功")
                messages = self.track_parse(check)
                for message in messages:
                    track_save = self.dispose(abnormal, self.modify_status(message))
                    save_data.append(track_save)
            else:
                while True:
                    print(f"{start_number}{end_number}数据缺失，再次请求")
                    time.sleep(random.randint(7, 10))
                    track_json = self.make_response(self.url, self.my_headers, data)
                    check = self.check_json_data(track_json)
                    if check != None:
                        print(f"{start_number}-{end_number}物流详情获取成功")
                        messages = self.track_parse(check)
                        for message in messages:
                            track_save = self.dispose(abnormal, self.modify_status(message))
                            save_data.append(track_save)
                        break
            start_number += 40
            end_number += 40
            totals -= 1
            time.sleep(random.randint(10, 15))
        self.save_track_status(save_data)


if __name__ == '__main__':
    with open('cookie.txt', 'r', encoding='utf-8') as fp:
        cookie = fp.read()
    track = Track_17(cookie)
    track.run()

