import time
import lackey

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PySide2.QtWidgets import QApplication, QMainWindow, QFileDialog


def google_map(driver):
    count = 1
    # driver.maximize_window()
    driver.get('https://www.google.com/maps/')
    time.sleep(2)
    for i in excel_read():
        print(' '.join(i[1:4]))
        try:
            wait.until(
                EC.presence_of_element_located((By.ID, "searchboxinput"))).send_keys(' '.join(i[1:4]))
            # driver.find_element_by_xpath('//*[@id="searchboxinput"]').send_keys(' '.join(i[1:4]))
            #             # time.sleep(0.5)
            driver.find_element_by_id("searchbox-searchbutton").click()
            time.sleep(1.5)
            try:
                post_code = driver.find_element_by_xpath('//div[@class="x3AX1-LfntMc-header-title-ma6Yeb-haAclf"]//h2[@jsinstance="0"]//span[1]').text
                print(post_code)
            except NoSuchElementException:
                print('在谷歌地图中没有找到对应邮编')
            time.sleep(1)
            lackey.click('D:\google_map\logo.png')
            time.sleep(1)
            lackey.click('D:\google_map\logo.png')
            time.sleep(1)
            print(str(i[0]) + ' ----------地址已查到!!!')
            print('\n')
        except lackey.Exceptions.FindFailed:
            count += 1
            print(count)
            excel_save(str(i[0]), count, wb)
            # print('chucuo')
            # # with open(filename_city_name, 'a', encoding='utf-8') as fp1 :
            # #     for i in city:
            # #         fp1.write(i)
            print(str(i[0]) + "-----不存在,订单号已存入xlsx")
            print('--------------')
            print('\n')
            continue
        finally:
            driver.find_element_by_xpath('//*[@id="searchboxinput"]').clear()
            time.sleep(1)


def excel_read():
    wb = load_workbook(choose)
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append(row)
    for x in range(1, len(rows)):
        data_text = []
        order = str(rows[x][1].value)
        city =  str(rows[x][10].value)
        mail_address_first = str(rows[x][11].value)
        mail_address_second = str(rows[x][12].value)
        data_text.append(order)
        data_text.append(city)
        data_text.append(mail_address_first)
        data_text.append(mail_address_second)
        item1 = data_text[0:4]
        # str_join = " ".join(item1)
        # str_join = str(rows[x][1].value) + str(rows[x][9].value) + str(rows[x][10].value) + str(rows[x][11].value)
        yield item1


def excel_save(order, i, wb):
    ws = wb.active
    ws.append(['订单号'])
    # ws.cell(row=2, column=1).value = get_column_letter(1)
    ws[f'A{i}'] = order
    file_location = 'D:\google_map\order.xlsx'
    wb.save(file_location)
    print(f'file location save to {file_location}')


def choose_file():
    app = QApplication([])
    MainWindow = QMainWindow()
    FileDialog = QFileDialog(MainWindow)
    FileDirectory = FileDialog.getOpenFileName(MainWindow, "标题")  # 选择目录，返回选中的路径
    # print(FileDirectory)
    # MainWindow.show()
    # app.exec_()
    print(f'当前读取的工作簿文件地址：{str(FileDirectory[0])}')
    return str(FileDirectory[0])


choose = choose_file()
options = Options()
options.add_argument('---disable-plugins')
options.add_argument('disable-infobars')  # 不显示Chrome正在受自动软件控制
options.add_argument("no-default-browser-check")
options.add_experimental_option('useAutomationExtension', False)
options.add_argument('disable-flash-core-animation')
driver = webdriver.Chrome(options=options)
wait = WebDriverWait(driver, 10)
wb = workbook.Workbook()
google_map(driver)


