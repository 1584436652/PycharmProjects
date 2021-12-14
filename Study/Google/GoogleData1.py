from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from openpyxl import workbook
from openpyxl import load_workbook
import time
import lackey


def google_map(driver, count):
    driver.get('https://www.google.com/maps/')
    time.sleep(2)
    print(count)
    for j in range(1, count):
        for i in excel_read():
            print()
            try:
                print(i)
                driver.find_element_by_xpath('//*[@id="searchboxinput"]').send_keys(i)
                driver.find_element_by_id("searchbox-searchbutton").click()
                time.sleep(2)
                lackey.doubleClick('./logo.png')
                time.sleep(2)
            except lackey.Exceptions.FindFailed:
                # print('chucuo')
                # # with open(filename_city_name, 'a', encoding='utf-8') as fp1 :
                # #     for i in city:
                # #         fp1.write(i)
                print("-----不存在,已保存")
                continue
            finally:
                driver.find_element_by_xpath('//*[@id="searchboxinput"]').clear()

def excel_read():
    wb = load_workbook('C:\\Users\\Administrator\\Desktop\\COD.xlsx')
    ws = wb.active
    rows = []
    for row in ws.iter_rows():
        rows.append(row)
    for x in range(1, len(rows)):
        data_text = []
        order = str(rows[x][1].value)
        city =  str(rows[x][9].value)
        mail_address_first = str(rows[x][10].value)
        mail_address_second = str(rows[x][11].value)
        data_text.append(order)
        data_text.append(city)
        data_text.append(mail_address_first)
        data_text.append(mail_address_second)
        item1 = data_text[1:4]
        # str_join = " ".join(item1)
        # str_join = str(rows[x][1].value) + str(rows[x][9].value) + str(rows[x][10].value) + str(rows[x][11].value)
        yield  item1
    return litem1

def excel_save():

   pass



if __name__ == '__main__':
    driver = webdriver.Chrome('D:\Download\chromedriver_win32\chromedriver.exe')
    # filename = 'C:\\Users\\Administrator\\Desktop\\city.txt'
    # filename_city_name = 'C:\\Users\\Administrator\\Desktop\\city_name.txt'
    # wb = workbook.Workbook()
    count = int(input('要处理的行数：'))
    google_map(driver, count)


