import re
import PySimpleGUI as sg

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import workbook
from openpyxl import load_workbook

# from PySide2.QtWidgets import QApplication, QMainWindow, QFileDialog


"""
腾伟订单地址核单
如果从表格里读取地址能在谷歌地图中匹配到邮编，则为正常
否则将没有匹配到邮编对应的订单号存入D:\google_map\order.xlsx
"""
class TenWei_Checklist(object):

    def __init__(self):
        self.get_url = 'https://www.google.com/maps/'
        self.wb = workbook.Workbook()

    # def tkinter_demo(self, h2_end):
    #     # 第1步，实例化object，建立窗口window
    #     window = tk.Tk()
    #     # 第2步，给窗口的可视化起名字
    #     window.title('腾伟核单')
    #     # 第3步，设定窗口的大小(长 * 宽)
    #     window.geometry('500x300')  # 这里的乘是小x
    #     # 在窗口界面设置放置Button按键
    #     b = tk.Button(window, text='start', font=('Arial', 12), width=10, height=1, command=self.google_map)
    #     b.pack()
    #     window.mainloop()

    # 读取表格地址到谷歌地图搜索
    def google_map(self):
        self.options = Options()
        # self.options.add_argument('--start-maximized')
        self.options.add_argument('disable-infobars')  # 不显示Chrome正在受自动软件控制
        self.driver = webdriver.Chrome(options=self.options)
        self.wait = WebDriverWait(self.driver, 3)
        count = 1
        self.driver.get(self.get_url)
        for i in self.excel_read():
            print(f'地址：{" ".join(i[1:4])}')
            try:
                self.wait.until(
                    EC.presence_of_element_located((By.ID, "searchboxinput"))).send_keys(' '.join(i[1:4]))
                self.wait.until(
                    EC.element_to_be_clickable((By.ID, 'searchbox-searchbutton'))).click()
                time.sleep(2)
                post_code = self.wait.until(
                    EC.presence_of_element_located((By.XPATH, '//div[@class="x3AX1-LfntMc-header-title-ij8cu"]'))
                )
                post_code_h2 = post_code.find_elements_by_tag_name('h2')
                item = []
                for h2 in post_code_h2:
                    item.append(h2.text)  # 获取当前h2标签文本添加到列表
                h2_splicing = ''.join(item)
                pattern = re.compile(r'\d+')
                extract = pattern.findall(h2_splicing)
                h2_end = ''.join(extract)
                print(h2_end)
                # print(len(h2_end))
                if len(h2_end) >= 4:
                    print(str(i[0]) + ' ----------地址已查到!!!')
                    print('\n')
                else:
                    count += 1
                    # print(count)
                    self.excel_save(str(i[0]), count)
                    print('在谷歌地图中没有找到对应邮编_first')
                    print(str(i[0]) + "-----订单号已存入xlsx")
                    print('\n')
                    continue
            except TimeoutException:
                count += 1
                # print(count)
                self.excel_save(str(i[0]), count)
                print('在谷歌地图中没有找到对应邮编_second')
                print(str(i[0]) + "-----订单号已存入xlsx")
                print('\n')
                continue
            finally:
                self.wait.until(
                    EC.element_to_be_clickable((By.ID, 'searchboxinput'))).clear()
        self.driver.close()

    # 读取表格数据
    def excel_read(self):
        wb = load_workbook(self.PySimpleGUI_choose_file())
        ws = wb.active
        rows = []
        for row in ws.iter_rows():
            rows.append(row)
        for x in range(1, len(rows)):
            data_text = []
            order = str(rows[x][1].value)
            city = str(rows[x][10].value)
            mail_address_first = str(rows[x][11].value)
            mail_address_second = str(rows[x][12].value)
            data_text.append(order)
            data_text.append(city)
            data_text.append(mail_address_first)
            data_text.append(mail_address_second)
            item1 = data_text[0:4]
            yield item1

    # 存储没有匹配到邮编对应的订单号
    def excel_save(self, order, i):
        ws = self.wb.active
        ws.append(['订单号'])
        # ws.cell(row=2, column=1).value = get_column_letter(1)
        ws[f'A{i}'] = order
        file_location = 'D:\Work_Code\Google\order.xlsx'
        self.wb.save(file_location)
        print(f'文件存储地址{file_location}')

    # 选择要读取的表格文件
    # def choose_file(self):
    #     app = QApplication([])
    #     MainWindow = QMainWindow()
    #     FileDialog = QFileDialog(MainWindow)
    #     FileDirectory = FileDialog.getOpenFileName(MainWindow, "标题")  # 选择目录，返回选中的路径
    #     # print(FileDirectory)
    #     # MainWindow.show()
    #     # app.exec_()
    #     # print(f'当前读取的工作簿文件地址：{str(FileDirectory[0])}')
    #     return str(FileDirectory[0])

    def PySimpleGUI_choose_file(self):
        file_address = sg.popup_get_file('请选择你要读取的表格文件：')
        return file_address

    def main(self):
        tw.google_map()


tw = TenWei_Checklist()
tw.main()




