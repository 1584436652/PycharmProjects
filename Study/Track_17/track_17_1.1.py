'''
@Time    : 2021/10/12 12:48
@Author  : LKT
@FileName: track_17_1.1.py
@Software: PyCharm
 
'''
import requests
import json
import math
import time
import random

from openpyxl import load_workbook
from retry import retry
"""
17track查单
"""

class Track_17(object):

    def __init__(self, cookie=None):
        self.headers = {
            'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)"
                          " Chrome/92.0.4515.159 Safari/537.36",
            'Referer': 'https://t.17track.net/zh-cn',
            'Cookie': cookie
        }
        self.proxies = {
            # "https": "https://47.108.154.194:8080"
        }
        self.url = 'https://t.17track.net/restapi/track'

    # 读取跟踪号表格
    def read_trcak(self, file_path):
        wb = load_workbook(file_path)
        ws = wb.active
        rows = []
        tracking_number_list = []
        for row in ws.iter_rows():
            rows.append(row)
        for x in range(1, len(rows)):
            # 物流跟踪号
            tracking_number = str(rows[x][0].value)
            # message = str(rows[x][1].value)
            # nation = str(rows[x][2].value)
            tracking_number_list.append(tracking_number)
        return tracking_number_list

    # 构造data
    def trcak_data(self, track_list: list, *args):
        messages= []
        for order in track_list[args[0]:args[1]]:
            result = {"num":order,"fc":0,"sc":0}
            messages.append(result)
        data = {
            "data": messages,
            "guid":"",
            "timeZoneOffset":-480
        }
        return data

    @retry(delay=200)
    def make_response(self, url, data):
        print(f'正在请求--{url}')
        response = requests.post(url=url, headers=self.headers, data=json.dumps(data), proxies=self.proxies)
        print(response.status_code)
        html_data = response.json()
        print(html_data)
        if html_data['dat']:
            return html_data
        print("您的查询过于频繁，请稍后查询，等待重新尝试中(200s)...")
        raise Exception

    # 解析json获取物流信息
    def track_parse(self, track_json):
        orders = track_json["dat"]
        for order in orders:
            # 跟踪号
            track_numbers = order["no"]
            # 物流状态
            delivery_status = str(order["track"]["e"])
            # print(track_numbers, delivery_status)
            # 物流详情
            delivery_details = order["track"]["z1"]
            track_append = []
            for delivery_detail in delivery_details:
                # 到达站点时间
                delivery_date = delivery_detail["a"]
                # 到达站点地址
                delivery_address = delivery_detail["z"]
                # print(delivery_date, delivery_address
                track_append.append({delivery_date: delivery_address})
                # 返回跟踪后全部物流详情
            yield {
                track_numbers: [delivery_status, track_append]
            }

    # 修改解析后的物流状态
    def modify_status(self, messages: dict):
        status = {
            "0": "查询不到",
            "10": "运输途中",
            "35": "投递失败",
            "30": "到达待取",
            "40": "成功签收",
            "50": "可能异常",
            "60": "运输过久",
        }
        for message_key, massage_value in messages.items():
            # 把对应的物流状态改为中文
            messages[message_key][0] = status[messages[message_key][0]]
        # print(messages)
        return messages

    # 读取异常对应中文
    def abnormal_detail(self, filename):
        with open(f'{filename}.json', 'r', encoding='utf-8') as f:
            abnormal_json = f.read()
            abnormal_list = json.loads(abnormal_json)
            # print(abnormal_list)
            return abnormal_list

    # 用abnormal_detail相对应的可能异常改为中文
    def dispose(self, lists: list, dicts: dict):
        end_save_to = []
        for i, j in dicts.items():
            end_save_to.append(i)
            end_save_to.append(j[0])
        for item in lists:
            for abnormal_key, abnormal_values in item.items():
                for address in dicts.values():
                    for titles in address[1]:
                        for date, title in titles.items():
                            if address[0] == "可能异常":
                                # 判断对应异常是否在地址详情里，查询到就改为中文
                                if abnormal_key in title:
                                   alter =  titles[date] = abnormal_values
                                   end_save_to.append(alter)
        print(end_save_to)

    def choice_may_be_abnormal(self, results):
        for address in dicts.values():
            if address[0] == "可能异常":
                for date, title in titles.items():
                    if abnormal_key in title:
                        titles[date] = abnormal_values

    def save_track_status(self):
        ws = self.wb.active
        ws.append(['订单号'])
        # ws.cell(row=2, column=1).value = get_column_letter(1)
        ws[f'A{i}'] = order
        file_location = 'D:\Work_Code\Google\order.xlsx'
        self.wb.save(file_location)
        print(f'文件存储地址{file_location}')

    # 跟踪号条数/40向上取整
    def ceil_int(self, totals):
        return math.ceil(totals / 40)

    def run(self):
        track_list = track.read_trcak(r'C:\Users\Administrator\Desktop\track.xlsx')
        totals = self.ceil_int(len(track_list))
        abnormal = self.abnormal_detail("匈牙利")
        # 一次只能查找40条
        start_number = 0
        end_number = 40
        while totals >= 1:
            data = self.trcak_data(track_list, start_number, end_number)
            track_json = self.make_response(self.url, data)
            messages = self.track_parse(track_json)
            for message in messages:
                self.dispose(abnormal, self.modify_status(message))
                print()

            start_number += 40
            end_number += 40
            totals -= 1
            time.sleep(random.randint(8, 11))

if __name__ == '__main__':
    with open('cookie.txt', 'r', encoding='utf-8') as fp:
        cookie = fp.read()
    track = Track_17(cookie)
    track.run()

